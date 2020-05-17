import time
from ctypes import windll
from random import random, randint
from tkinter import *
import openpyxl

# Подключаем классы из NNet.py
from NNet import *
from Player import *


def paint(net):
    # Ищем самый жирный слой (в котором больше всего нейронов)
    canvas.delete("All")
    maxN = 0
    for i in range(len(netP)):
        if len(netP[i]) > len(netP[maxN]):
            maxN = i

    # Расчитываем размер под один нейрон
    sizeNY = (windll.user32.GetSystemMetrics(1) - 200) / len(netP[maxN])
    sizeNX = (windll.user32.GetSystemMetrics(0) - 200) / len(net.net)

    # Проходимся по нейронам и отрисовываем их выходы
    for i in range(len(net.net) - 1):
        ay = (windll.user32.GetSystemMetrics(1) - len(net.net[i]) * sizeNY) / 2
        by = (windll.user32.GetSystemMetrics(1) - len(net.net[i + 1]) * sizeNY) / 2
        ax = (windll.user32.GetSystemMetrics(0) - len(net.net) * sizeNX) / 2
        for j in range(len(net.net[i])):
            for k in range(len(net.net[i][j].exits)):
                canvas.create_line(ax + i * sizeNX + sizeNX * 0.4, ay + j * sizeNY + sizeNY * 0.4,
                                   ax + (i + 1) * sizeNX + sizeNX * 0.4,
                                   by + net.net[i][j].exits[k] * sizeNY + sizeNY * 0.4)

    # Поверх выходов рисуем нейроны
    for i in range(len(net.net)):
        ay = (windll.user32.GetSystemMetrics(1) - len(net.net[i]) * sizeNY) / 2
        ax = (windll.user32.GetSystemMetrics(0) - len(net.net) * sizeNX) / 2
        for j in range(len(net.net[i])):
            canvas.create_oval(ax + i * sizeNX, ay + j * sizeNY, ax + i * sizeNX + sizeNX * 0.8,
                               ay + j * sizeNY + sizeNY * 0.8, fill='pink')
            # И их значения
            canvas.create_text(ax + i * sizeNX + sizeNX * 0.4, ay + j * sizeNY + sizeNY * 0.4,
                               text=round(net.net[i][j].value, 2))
    # Пререрисовываем окно
    window.update()


# Создаем окно
window = Tk()
# Создаем то, где будем рисовать
canvas = Canvas(window, width=windll.user32.GetSystemMetrics(0), height=windll.user32.GetSystemMetrics(1))
# Соединяем это воедино
canvas.pack()

players = []
for k in range(20):
    d = {}
    #  Создаем структуру нейронной сети
    netS = [37, randint(3, 10), 6, 0]
    netP = []
    for j in range(len(netS) - 1):
        a = []
        for s in range(netS[j]):
            if j != 0:
                f = []
                for i in range(netS[j - 1]):
                    f.append(i)
                cash = int(randint(2, netS[j - 1]) / 2)
                for i in range(cash):
                    c = randint(0, len(f) - 1)
                    d.setdefault(f[c], [0, 1])
                    f.pop(c)
            a.append(Perceptron(d, []))
            d.clear()
        netP.append(a)
        d.clear()

    for j in range(len(netS) - 1):
        for s in range(netS[j]):
            for l in range(netS[j + 1]):
                for g in netP[j + 1][l].enters.keys():
                    if g == s:
                        netP[j][s].exits.append(l)

    # Создаем сеть опираясь на структуру
    netL = NNet(netP)
    players.append(Player(netL))
game = Game(players)
game.weBrokeThisGame(15000, 30, 100)
# Отрисовываем
# f = []
# for i in range(36):
#     f.append(randint(0, 6))
# netL.giveEnters(f)
# netL.update()
# canvas.create_text(50, 50, text="Нарисовали")
# # paint(netL)
#
# excel = openpyxl.load_workbook(filename="excel/xl.xlsx")
# sheet = excel.create_sheet('net')д
# print(sheet['G1'].value)
# sheet['L1'] = netL.getSolution()
# for i in range(len(netL.net)):
#     for j in range(len(netL.net[i])):
#         for k in netL.net[i][j].enters:
#             sheet.cell(column=(i * 4 + 1), row=(j * 38 + 1 + k)).value = f"={chr(ord('A') + (i - 1) * 4 + 2)}{1 + k * 38}"
#             sheet.cell(column=(i * 4 + 2), row=(j * 38 + 1 + k)).value = netL.net[i][j].enters[k][1]
#         if i == 0:
#             sheet.cell(column=(i * 4 + 3), row=(j * 38 + 1)).value = netL.net[i][j].value
#         else:
#             print(f"=SUMPRODUCT({chr(ord('A') + i * 4)}{j * 38 + 1}:{chr(ord('A') + i * 4)}{j * 38 + 37},{chr(ord('A') + i * 4 + 1)}{j * 38 + 1}:{chr(ord('A') + i * 4 + 1)}{j * 38 + 37})/COUNTIF({chr(ord('A') + i * 4 + 1)}{j * 38 + 1}:{chr(ord('A') + i * 4 + 1)}{j * 38 + 37},\">0\")")
#             sheet.cell(column=(i*4 + 3), row=(j*38+1)).value = f"=SUMPRODUCT({chr(ord('A') + i * 4)}{j * 38 + 1}:{chr(ord('A') + i * 4)}{j * 38 + 37},{chr(ord('A') + i * 4 + 1)}{j * 38 + 1}:{chr(ord('A') + i * 4 + 1)}{j * 38 + 37})/(COUNTIF({chr(ord('A') + i * 4 + 1)}{j * 38 + 1}:{chr(ord('A') + i * 4 + 1)}{j * 38 + 37},\">0\")+1)"
#         # sheet.cell(column=(i*4 + 4), row=(j*38+1)).value = f"=SUMPRODUCT(B37:B41,C37:C41)"
#
#
# for i in range(6):
#     print(netL.net[len(netL.net)-1][i].value, end="\t")
# print()
# # time.sleep(5)
# canvas.create_rectangle(0, 0, windll.user32.GetSystemMetrics(0), windll.user32.GetSystemMetrics(1), fill="white")
# print("Обучение")
# teacher = Teacher(netL).updateSumWay(netL.net)
# for i in range(len(teacher.sumWay)):
#     print(teacher.sumWay[i])
# netL.learn(False, netL.getSolution(), teacher)
# sheet = excel.create_sheet('netLearn')
# sheet['L1'] = netL.getSolution()
# for i in range(len(netL.net)):
#     for j in range(len(netL.net[i])):
#         for k in netL.net[i][j].enters:
#             sheet.cell(column=(i * 4 + 1), row=(j * 38 + 1 + k)).value = f"={chr(ord('A') + (i - 1) * 4 + 2)}{1 + k * 38}"
#             sheet.cell(column=(i * 4 + 2), row=(j * 38 + 1 + k)).value = netL.net[i][j].enters[k][1]
#         if i == 0:
#             sheet.cell(column=(i * 4 + 3), row=(j * 38 + 1)).value = netL.net[i][j].value
#         else:
#             sheet.cell(column=(i*4 + 3), row=(j*38+1)).value = f"=SUMPRODUCT({chr(ord('A') + i * 4)}{j * 38 + 1}:{chr(ord('A') + i * 4)}{j * 38 + 37},{chr(ord('A') + i * 4 + 1)}{j * 38 + 1}:{chr(ord('A') + i * 4 + 1)}{j * 38 + 37})/(COUNTIF({chr(ord('A') + i * 4 + 1)}{j * 38 + 1}:{chr(ord('A') + i * 4 + 1)}{j * 38 + 37},\">0\")+1)"
#         # sheet.cell(column=(i*4 + 4), row=(j*38+1)).value = f"=SUMPRODUCT(B37:B41,C37:C41)"
# excel.save('excel/xl.xlsx')
#
# print(netL.net[len(netL.net)-1][0].value)
# print(netL.net[len(netL.net)-1][0].enters.values())
#
# netL.giveEnters(f)
# netL.update()
# canvas.create_text(50, 50, text="Обученно")
# paint(netL)
# for i in range(6):
#     print(netL.net[len(netL.net)-1][i].value, end="\t")
# time.sleep(5)
# canvas.create_rectangle(0, 0, windll.user32.GetSystemMetrics(0), windll.user32.GetSystemMetrics(1), fill="white")
# print("Мутация")
# netL.mutate(0.5)
# netL.giveEnters(f)
# netL.update()
# canvas.create_text(50, 50, text="Перерисовано")
# paint(netL)
# time.sleep(3)
# net.mutate(0.1)
# # Отрисовываем
# paint()
# # Задерживаем окно на экране
# window.mainloop()
