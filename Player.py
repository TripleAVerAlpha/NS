import copy
from random import randint
import openpyxl
import time
from NNet import *


class Player:
    def __init__(self, net):
        self.money = 10
        self.maxMoney = 0
        self.net = NNet(net.net)
        self.answer = -1


class Game:
    def __init__(self, players):
        self.players = players
        self.score = [0] * len(self.players)
        self.game = [0] * 36
        self.tour = 0
        self.beginTime = time.time()

    def weBrokeThisGame(self, learnTour, gameTour, allTour):
        for i in range(allTour):
            print(f"Популяция {i+1} Средне время на популяцию {(time.time() - self.beginTime)/(i+0.0000001)}")
            timeNow = (time.time() - self.beginTime)/(i+0.0000001) * (allTour - i)
            print(f"Осталось времени:{int(timeNow / (3600 * 24))} {int(timeNow % (3600 * 24) / 3600)} часов {int(timeNow % 3600 / 60)} минут {timeNow % 60} секунд")
            # Готовим массив под учителей
            teacher = []
            for g in range(len(self.players)):
                # Создаем учителей
                teacher.append(Teacher(self.players[g].net))
            for j in range(learnTour):
                a = time.time()
                self.simulatedGame(teacher)
                timeNow = timeNow - (time.time() - a)
                print(f"\r Идет {i + 1}/{allTour} тур, {j+1}/{learnTour}(Обучение) \n Затраченое время:{time.time() - a}, Осталось времени: {int(timeNow / (3600 * 24))} дней {int(timeNow % (3600 * 24) / 3600)} часов {int(timeNow % 3600 / 60)} минут {int(timeNow % 60)} секунд", end="")
            print()
            for k in range(gameTour):
                print(f"\r Идет {i + 1}/{allTour} тур, {k+1}/{gameTour}", end="")
                self.simulatedGame(None)
            print()
            self.savePlayers(teacher, gameTour)
            self.tour += 1
            # Итак, игры первой группы нейро сетей закончены, самое время создать новую группу
            d = {}
            for k in range(len(self.players)):
                d.setdefault(k, self.score[k])
            playerCash = [None] * 20
            listD = list(d.items())
            listD.sort(key=lambda i: i[1])
            for j in range(1, 6):
                playerCash[listD[len(self.players) - j][0]] = Player(self.players[listD[len(self.players) - j][0]].net)
            for j in range(20):
                if playerCash[j] is None:
                    playerCash[j] = Player(self.players[listD[len(self.players) - (j % 5 + 1)][0]].net.mutate(0.1))
            self.players = list(playerCash)
            # Опираясь на Очки выбираем 5 лучших нс
            # Кладем их и 2 копии в новый массив playerCash
            # У копий вызываем метод mutate
            # list(playerCash) приравниваем players
            timeNow = time.time() - self.beginTime
            print(f"Потрачено времени: {int(timeNow / (3600 * 24))} дней {int(timeNow % (3600 * 24) / 3600)} часов {int(timeNow % 3600 / 60)} минут {int(timeNow % 60)} секунд")

    def simulatedGame(self, teacher):
        self.game = [0] * 36
        # Выдали стартовый набор инвестора
        for l in range(len(self.players)):
            self.players[l].money = 10

        # Начинаем игру
        for h in range(6):
            # print("Тур")
            #  Рассказали участиникам результаты прошлых игр
            for l in range(len(self.players)):
                self.players[l].net.giveEnters(self.game)
            # Приняли от них ответы и создали список результатов на данный тур
            for l in range(len(self.players)):
                self.players[l].answer = self.players[l].net.getSolution() + h * 6
                self.game[self.players[l].answer] += 1
            # У нас есть старт ап готовим для него число от 1 до 6
            startUpNumber = randint(1, 6)
            # Для МММ готовим варианты их доходности, Ключ = Кол-во вкладчиков, Значение = дохлдность
            mMm = {1: 1.5, 2: 2, 3: 3, 4: 6}
            # Высчитываем доходность
            m = [0] * 6
            # Сбер
            m[0] = 1.1
            # Газпром
            m[1] = (10 / self.game[self.players[l].answer])
            # Яндекс
            m[2] = (5 / self.game[self.players[l].answer])
            # ГазпромНефть
            m[3] = (self.game[self.players[l].answer - 2] / self.game[self.players[l].answer])
            # Старт ап
            if startUpNumber == 6:
                m[4] = 7
            else:
                m[4] = 0.8
            # мМм
            try:
                m[5] = mMm[self.game[self.players[l].answer]]
            except KeyError as e:
                m[5] = 0.3
            # Выдаем игрокам их кровные денюжки
            for l in range(len(self.players)):
                # print(f"Игрок {l} заработал {(m[self.players[l].answer - h * 6] - 1)*100}% у него было {self.players[l].money} стало {self.players[l].money * m[self.players[l].answer - h * 6]}")
                self.players[l].money = self.players[l].money * m[self.players[l].answer - h * 6]

            # Если учителя нет, значит раунд без обучения
            if teacher is not None:
                # Самое время обучаться
                # Берем ищем максимальную доходность из массива m
                maxM = 0
                for i in range(len(m)):
                    if m[i] > m[maxM]:
                        maxM = i
                # Проходим по всем игрокам
                # Если их ответ - h * 6 равен максимальному ответу, то вызываем функцию обучения и говорим что ответ верен
                # Иначе передаем в функцию, то что ответ не верен
                for i in range(len(self.players)):
                    if self.players[i].answer - h * 6 == maxM:
                        self.players[i].net.learn(True, self.players[i].answer - h * 6, teacher[i])
                    else:
                        self.players[i].net.learn(False, self.players[i].answer - h * 6, teacher[i])
            else:
                if h == 5:
                    d = {}
                    d1 = {0: 10, 1: 8, 2: 5, 3: 2, 4: 1}
                    for i in range(len(self.players)):
                        d.setdefault(i, self.players[i].money)
                    listD = list(d.items())
                    listD.sort(key=lambda i: i[1])
                    for i in range(1, 6):
                        # print(self.score)
                        # print(len(self.players))
                        # print(len(self.score))
                        # print(d1[i - 1])
                        self.score[listD[len(self.players) - i][0]] += d1[i - 1]

                    # Если учителя нет, добавляем Очки в зависимости от места:
                    # 1 - 10 очков
                    # 2 - 8
                    # 3 - 5
                    # 4 - 2
                    # 5 - 1
                    # Остальные - 0
        for i in range(len(self.players)):
                self.players[i].maxMoney += self.players[i].money
        # a = ["Сбербанк    ", "Газпром      ", "Яндекс       ", "ГазпромНефть", "СтартUp      ", "мМм         "]
        # for i in range(6):
        #     print(a[i], end='\t')
        #     for j in range(6):
        #         print(self.game[i + j * 6], end='\t')
        #     print()
        # print(self.game)

    def savePlayers(self, teacher, kolTure):
        excel = openpyxl.load_workbook(filename='excel/xl.xlsx')
        sheet = excel['players']
        for i in range(len(self.players)):
            neyronKol = 0
            pathKol = 0
            for j in range(len(self.players[i].net.net)):
                neyronKol += len(self.players[i].net.net[j])
            for j in range(len(self.players[i].net.net[0])):
                for k in range(6):
                    pathKol += teacher[i].sumWay[0][j][k]
            sheet.cell(row=(1 + i), column=(1 + self.tour)).value = self.players[i].maxMoney / kolTure
        sheet = excel['game']
        for i in range(len(self.game)):
            sheet.cell(row=(1 + i), column=(1 + self.tour)).value = self.game[i]
        excel.save('excel/xl.xlsx')
